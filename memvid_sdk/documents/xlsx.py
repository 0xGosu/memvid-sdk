"""Excel Parser — two strategies:

- ``parse_xlsx()`` — Python-based via openpyxl (legacy, per-sheet flat text)
- ``parse_xlsx_structured()`` — Rust-native via memvid-core (high accuracy,
  header-value pairing, merged cell support, semantic chunking)
"""

from __future__ import annotations

import os
import time
from typing import Optional, Any, Dict, List

from . import ParseOptions, ParseResult, DocumentItem


def get_cell_value(cell_value: Any) -> Optional[str]:
    """Extract the display value from a cell, handling formulas and special types."""
    if cell_value is None:
        return None

    # Handle formula results (openpyxl stores computed values directly)
    # Just convert to string
    if isinstance(cell_value, (int, float)):
        return str(cell_value)
    elif isinstance(cell_value, str):
        return cell_value if cell_value.strip() else None
    elif hasattr(cell_value, "isoformat"):  # datetime
        return cell_value.isoformat().split("T")[0]
    else:
        return str(cell_value) if cell_value else None


def parse_xlsx(file_path: str, options: Optional[ParseOptions] = None) -> ParseResult:
    """
    Parse an Excel file, extracting text per sheet.

    Args:
        file_path: Path to the Excel file
        options: Parsing options (max_items limits sheets)

    Returns:
        ParseResult with per-sheet items
    """
    filename = os.path.basename(file_path)

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel parsing. "
            "Install with: pip install openpyxl"
        )

    try:
        workbook = load_workbook(file_path, data_only=True)  # data_only=True gets computed values
        sheet_names = workbook.sheetnames
        max_items = (options or {}).get("max_items") or len(sheet_names)

        items: list[DocumentItem] = []
        for i, sheet_name in enumerate(sheet_names[:max_items]):
            sheet = workbook[sheet_name]
            text_lines: list[str] = []

            for row in sheet.iter_rows():
                values: list[str] = []
                for cell in row:
                    value = get_cell_value(cell.value)
                    if value:
                        values.append(value)
                if values:
                    text_lines.append(" | ".join(values))

            if text_lines:
                items.append({
                    "number": i,
                    "name": sheet_name,
                    "text": "\n".join(text_lines),
                })

        workbook.close()

        return {
            "type": "xlsx",
            "filename": filename,
            "total_items": len(sheet_names),
            "items": items,
        }
    except Exception as e:
        raise RuntimeError(
            f'Failed to parse Excel file "{filename}": {e}. '
            f"Ensure the file is a valid .xlsx/.xls file."
        )


def parse_xlsx_structured(
    file_path: str,
    *,
    max_chars: int = 1200,
    max_chunks: int = 500,
) -> Dict[str, Any]:
    """Parse an XLSX file using the Rust structured extraction pipeline.

    This provides much higher search accuracy than ``parse_xlsx()`` by:

    - Detecting table boundaries and headers automatically
    - Formatting rows as ``Header: Value | Header: Value`` pairs
    - Propagating merged cells
    - Detecting number formats (dates, currency, percentages)
    - Never splitting rows across chunk boundaries

    Args:
        file_path: Path to the XLSX file.
        max_chars: Target chunk size in characters (default: 1200).
        max_chunks: Maximum number of chunks to produce (default: 500).

    Returns:
        dict with keys:
            - ``text`` (str): Backward-compatible flat text.
            - ``chunks`` (list[dict]): Semantic chunks with ``text``,
              ``chunk_type``, ``index``, ``element_id``, ``context``.
            - ``tables`` (list[dict]): Detected tables with ``name``,
              ``sheet_name``, ``headers``, ``confidence``, ``column_types``, etc.
            - ``diagnostics`` (dict): Warnings and processing stats.
            - ``timing_ms`` (int): Extraction time in milliseconds.

    Example::

        result = parse_xlsx_structured("proforma.xlsx")
        print(f"{len(result['tables'])} tables, {len(result['chunks'])} chunks")

        # Ingest chunks into memvid for high-accuracy search
        for chunk in result["chunks"]:
            mv.put(title=f"Chunk {chunk['index']}", text=chunk["text"])
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    # Try native Rust extraction
    try:
        from .._lib import parse_xlsx_structured as _native_parse
        return _native_parse(
            file_path,
            max_chars=max_chars,
            max_chunks=max_chunks,
        )
    except ImportError:
        pass

    # Fallback: wrap legacy parse_xlsx output in structured format
    t0 = time.monotonic()
    legacy = parse_xlsx(file_path)
    elapsed_ms = int((time.monotonic() - t0) * 1000)

    chunks: List[Dict[str, Any]] = []
    for i, item in enumerate(legacy["items"]):
        sheet_label = item.get("name") or f"Sheet{item['number']}"
        chunks.append({
            "text": f"[Sheet: {sheet_label}]\n{item['text']}",
            "chunk_type": "Table",
            "index": i,
            "element_id": None,
            "context": None,
        })

    return {
        "text": "\n\n".join(item["text"] for item in legacy["items"]),
        "chunks": chunks,
        "tables": [],
        "diagnostics": {
            "warnings": ["Native structured extraction unavailable; using openpyxl fallback"],
            "tables_processed": 0,
            "tables_split": 0,
        },
        "timing_ms": elapsed_ms,
    }
